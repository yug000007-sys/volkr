import os
import re
from datetime import datetime
from pathlib import Path
from ftplib import FTP
import pdfplumber
from openpyxl import Workbook, load_workbook

# =========================
# CONFIG
# =========================
PDF_FOLDER = r"./pdfs"                 # folder with your Voelker PDFs
EXCEL_PATH = r"./voelker_output.xlsx"  # output excel
SHEET_NAME = "Data"

# Your Excel headers (custom mapping)
HEADERS = [
    "PDFName",
    "ReferralManager",   # <- Salesperson
    "QuoteNumber",
    "QuoteDate",
    "Company",           # <- ShipToName
    "Address",           # <- ShipToStreet
    "City",
    "State",
    "Zip",
    "TotalSales",        # <- QuoteTotal
    "Created_By",        # <- QuotedBy
    "CustomerNumber"
]

# FTP details (use env vars in real use)
FTP_HOST = os.getenv("FTP_HOST", "ftp.yourserver.com")
FTP_USER = os.getenv("FTP_USER", "username")
FTP_PASS = os.getenv("FTP_PASS", "password")
FTP_REMOTE_DIR = os.getenv("FTP_REMOTE_DIR", "/uploads/voelker")

UPLOAD_EXCEL_TOO = True


# =========================
# PDF PARSING HELPERS
# =========================
def extract_text_from_pdf(pdf_path: str) -> str:
    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            text_parts.append(t)
    return "\n".join(text_parts)

def find_first(pattern: str, text: str, flags=0) -> str:
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else ""

def parse_ship_to(text: str):
    """
    Attempts to extract Ship To name, street, city, state, zip
    based on Voelker layout where 'Ship To' section repeats address lines.
    """
    # Strategy:
    # Find the first block after "Ship To" that looks like:
    # NAME \n STREET \n CITY STATE ZIP
    # (Voelker sample repeats Sold To and Ship To; we just take the first good match.)
    ship_block = ""

    # Grab a slice starting at "Ship To" label if present
    idx = text.find("Ship To")
    slice_text = text[idx: idx + 800] if idx != -1 else text

    # Match address block:
    # line1: company
    # line2: street
    # line3: city state zip
    addr_match = re.search(
        r"Ship To\s*\n([A-Z0-9 &\.\-/,]+)\n([A-Z0-9 \.\-/,#]+)\n([A-Z \.\-']+)\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)",
        slice_text,
        re.MULTILINE
    )

    if addr_match:
        name = addr_match.group(1).strip()
        street = addr_match.group(2).strip()
        city = addr_match.group(3).strip()
        state = addr_match.group(4).strip()
        zipc = addr_match.group(5).strip()
        return name, street, city, state, zipc

    # Fallback: sometimes "Ship To" not captured in slice; try anywhere
    addr_match2 = re.search(
        r"\n([A-Z0-9 &\.\-/,]+)\n([A-Z0-9 \.\-/,#]+)\n([A-Z \.\-']+)\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)\n",
        text,
        re.MULTILINE
    )
    if addr_match2:
        return (
            addr_match2.group(1).strip(),
            addr_match2.group(2).strip(),
            addr_match2.group(3).strip(),
            addr_match2.group(4).strip(),
            addr_match2.group(5).strip(),
        )

    return "", "", "", "", ""

def parse_voelker_fields(text: str) -> dict:
    # These patterns rely on the header row style in your samples.
    quote_date = find_first(r"\n(\d{1,2}/\d{1,2}/\d{2})\s+\d{1,2}/\d{1,2}/\d{2}\s+", text)
    # Quote # looks like 01/125785
    quote_number = find_first(r"\n(0\d/\d{6})\s", text)
    quoted_by = find_first(r"\n(0\d/\d{6})\s+([A-Z][A-Za-z]+(?:\s+[A-Z][A-Za-z]+)*)\s+UPS", text)
    if not quoted_by:
        # fallback: try find after Quote # line
        quoted_by = find_first(r"\n0\d/\d{6}\s+([A-Z][A-Za-z]+(?:\s+[A-Z][A-Za-z]+)*)\s+", text)

    # Salesperson and Cust # appear on same header line in sample:
    # 9/24/25 12/31/25 JULIA CAO DAYTON 10980 NET 30 DAYS
    salesperson = ""
    cust_no = ""
    header_line = find_first(r"\n\d{1,2}/\d{1,2}/\d{2}\s+\d{1,2}/\d{1,2}/\d{2}\s+.+\n", text)
    if header_line:
        # Split line and infer last numeric token before Terms
        tokens = header_line.strip().split()
        # heuristic: cust # is first all-digit token of length>=4
        for i, tok in enumerate(tokens):
            if tok.isdigit() and len(tok) >= 4:
                cust_no = tok
                # salesperson is token immediately before cust (in your sample it's "DAYTON")
                if i - 1 >= 0:
                    salesperson = tokens[i - 1]
                break

    # Quote Total near bottom: "Quote Total" then amount line
    # Sample: Quote Total 23,540.00
    quote_total = find_first(r"Quote Total\s*\n?\s*([\d,]+\.\d{2})", text)

    ship_name, ship_street, ship_city, ship_state, ship_zip = parse_ship_to(text)

    return {
        "ReferralManager": salesperson,
        "CustomerNumber": cust_no,
        "QuoteDate": quote_date,
        "QuoteNumber": quote_number,
        "Created_By": quoted_by,
        "Company": ship_name,
        "Address": ship_street,
        "City": ship_city,
        "State": ship_state,
        "Zip": ship_zip,
        "TotalSales": quote_total.replace(",", "") if quote_total else ""
    }


# =========================
# EXCEL HELPERS
# =========================
def ensure_workbook(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
    # ensure headers present if empty
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(HEADERS)
    # ensure first row is headers
    if [c.value for c in ws[1]] != HEADERS:
        # If file exists but headers differ, don't overwrite silently
        raise ValueError(f"Excel headers mismatch. Expected {HEADERS}, found {[c.value for c in ws[1]]}")
    return wb, ws

def append_row(ws, row_dict: dict, pdf_name: str):
    row = []
    for h in HEADERS:
        if h == "PDFName":
            row.append(pdf_name)
        else:
            row.append(row_dict.get(h, ""))
    ws.append(row)


# =========================
# FTP UPLOAD
# =========================
def ftp_upload_files(file_paths, remote_dir, also_upload_excel=None):
    with FTP(FTP_HOST) as ftp:
        ftp.login(FTP_USER, FTP_PASS)
        # try change/create dir
        try:
            ftp.cwd(remote_dir)
        except Exception:
            # create nested dirs
            parts = [p for p in remote_dir.split("/") if p]
            cur = ""
            for p in parts:
                cur += f"/{p}"
                try:
                    ftp.cwd(cur)
                except Exception:
                    ftp.mkd(cur)
                    ftp.cwd(cur)

        for fp in file_paths:
            fn = os.path.basename(fp)
            with open(fp, "rb") as f:
                ftp.storbinary(f"STOR {fn}", f)

        if also_upload_excel and os.path.exists(also_upload_excel):
            fn = os.path.basename(also_upload_excel)
            with open(also_upload_excel, "rb") as f:
                ftp.storbinary(f"STOR {fn}", f)


def main():
    pdf_dir = Path(PDF_FOLDER)
    pdf_files = sorted([str(p) for p in pdf_dir.glob("*.pdf")])

    if not pdf_files:
        print(f"No PDFs found in {PDF_FOLDER}")
        return

    wb, ws = ensure_workbook(EXCEL_PATH)

    for pdf_path in pdf_files:
        pdf_name = os.path.basename(pdf_path)
        text = extract_text_from_pdf(pdf_path)
        data = parse_voelker_fields(text)
        append_row(ws, data, pdf_name)
        print(f"Parsed: {pdf_name} -> QuoteNumber={data.get('QuoteNumber')} TotalSales={data.get('TotalSales')}")

    wb.save(EXCEL_PATH)
    print(f"Saved Excel: {EXCEL_PATH}")

    # Upload PDFs (and optionally Excel) to FTP
    ftp_upload_files(
        pdf_files,
        FTP_REMOTE_DIR,
        also_upload_excel=EXCEL_PATH if UPLOAD_EXCEL_TOO else None
    )
    print(f"Uploaded {len(pdf_files)} PDFs to FTP: {FTP_REMOTE_DIR}")
    if UPLOAD_EXCEL_TOO:
        print("Uploaded Excel too.")

if __name__ == "__main__":
    main()
